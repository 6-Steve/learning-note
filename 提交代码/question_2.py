import math
import numpy as np
from scipy.optimize import fsolve
import openpyxl

# 打开Excel文件
workbook = openpyxl.load_workbook(r'F:\res_test\result2.xlsx')
sheet_1 = workbook['Sheet1']

def equation_1(theta1, b, theta0, L):
    term1 = 0.5 * theta0 * np.sqrt(1 + theta0 ** 2) + 0.5 * np.log(theta0 + np.sqrt(1 + theta0 ** 2))
    term2 = 0.5 * theta1 * np.sqrt(1 + theta1 ** 2) + 0.5 * np.log(theta1 + np.sqrt(1 + theta1 ** 2))
    return b * term1 - b * term2 - L
#切片赋零值
def set_to_zero_slice(lst, x, y):
    lst[x:y + 1] = [0] * (y - x + 1)  # 使用切片赋值，y-x+1计算切片长度
    return lst

def get_x(theta):
    return b * theta * math.cos(theta)

def get_y(theta):
    return b * theta * math.sin(theta)

def get_k(theta):
    return (math.sin(theta) + theta * math.cos(theta)) / (math.cos(theta) - theta * math.sin(theta))

def get_cos_alpha(k, k12):
    # 其中k为k1或k2,返回cosα 1或2
    return abs((1 + k * k12) / math.sqrt((1 + k ** 2) * (1 + k12 ** 2)))

# 求解龙头把手后的第一个把手(求解theta)  （递推）
def equation(theta, b, theta1, longth):
    # theta1为前一个，theta为当前所求
    return b ** 2 * (theta1 ** 2 + theta ** 2 - 2 * theta1 * theta * (
                math.cos(theta1) * math.cos(theta) + math.sin(theta1) * math.sin(theta))) - longth ** 2

def get_alpha(k):
    res_a=math.atan(k)
    # 如果为-pi/2到0
    if res_a<0:
        return math.pi+res_a
    else:
        return res_a
    
# 参数以 米 为单位
theta0 = 32 * np.pi
v1 = 1.0
b = 0.55/ (2 * np.pi)#``````````````````````````````````````````````````````````
longth = 2.86
longth_2 = 1.65

# 改良的搜索算法参数，用于求解高精度解
pre_t=419.5845
las_t=419.5865
foot=0.0000001

# 初始化一个列表来存储解
solutions = []

# 循环L从1到300
# 求的是L从1到300时  龙头第一个把手的theta
for L_value in np.arange(pre_t, las_t, foot):
    # 初始猜测值（可能需要根据实际情况调整）
    initial_guess = 20 * np.pi  
    try:
        solution = fsolve(equation_1, initial_guess, args=(b, theta0, L_value))
        if np.all(np.isfinite(solution)):  # 确保解是有限数
            solutions.append(solution[0])  # 将解添加到列表中
        else:
            solutions.append(np.nan)  # 如果解不是有限数，则记录为NaN
    except Exception as e:
        print(f"在L={L_value}时求解失败: {e}")
        solutions.append(np.nan)  # 捕获异常并记录为NaN

#以π为单位输出解
solutions_in_pi = [sol / np.pi for sol in solutions]

# print(len(solutions_in_pi)) 输出可得大小为300
# 故从编号为0到299
# 这里 i 表示数组solutions_in_pi的编号 
for i in range(0, len(solutions_in_pi)):
    # 初始化阶段：
    # 龙头第一个把手的theta
    theta1 = solutions_in_pi[i] * np.pi
    # 初始化列表来保存结果  (并且放这里起到 清零的作用)
    x_values = []
    y_values = []
    v_values = []
    theta_values = []

    initial_guess = theta1 + 0.5 * np.pi  # 初始猜测值，可以根据具体情况调整
    # 求解
    solutions = fsolve(equation, initial_guess, args=(b, theta1, longth))
    # 只要第一个解（最靠近初始猜测解的）
    res = solutions[0]
    # 输出(pi为为单位)
    theta_value = res / np.pi
    #求位置
    x1 = get_x(theta1)
    y1 = get_y(theta1)
    v1 = 1.0
    # 求速度
    if theta_value < 32:
        x2 = get_x(res)
        y2 = get_y(res)
        k1 = get_k(theta1)
        k2 = get_k(res)
        k12 = (y2 - y1) / (x2 - x1)
        cos_1 = get_cos_alpha(k1, k12)
        cos_2 = get_cos_alpha(k2, k12)
        v2 = v1 * cos_1 / cos_2
    else:
        x2 = 0.0
        y2 = 0.0
        v2 = 0.0
        theta_value = 0.0
    x_values.append(x1)
    y_values.append(y1)
    v_values.append(v1)

    x_values.append(x2)
    y_values.append(y2)
    v_values.append(v2)
    theta_values.append(theta_value)

    theta1 = res

    # 设置碰撞标识
    flag=0
    num=0
    ### 进行 碰撞 判断
    # 求龙头A1 A2 的x y
    x__1=x_values[0]
    y__1=y_values[0]
    x__2=x_values[1]
    y__2=y_values[1]
    k_a1_a2=(y__2-y__1)/(x__2-x__1)
    # 求alpha_12 弧度制
    alpha_12=get_alpha(k_a1_a2)
    # 求用于碰撞判断的x,y
    c_x_1=x__1+0.275*math.cos(alpha_12)-0.15*math.sin(alpha_12)
    c_y_1=y__1+0.275*math.sin(alpha_12)+0.15*math.cos(alpha_12)

    c_x_2=x__2-0.275*math.cos(alpha_12)-0.15*math.sin(alpha_12)
    c_y_2=y__2-0.275*math.sin(alpha_12)+0.15*math.cos(alpha_12)

    for j in range(222):
        if theta1 == 0.0:
            # 该 时刻 剩余的板凳数据全为 全为0
            set_to_zero_slice(x_values, j + 2, 223)
            set_to_zero_slice(y_values, j + 2, 223)
            set_to_zero_slice(v_values, j + 2, 223)
            set_to_zero_slice(theta_values, j + 2, 223)
            # 之后的 时刻 对应所有板凳的全都为空了（不再进行操作）
            break
        # 计算初始猜测值
        initial_guess = theta1 + 0.5 * np.pi
        solutions_2 = fsolve(equation, initial_guess, args=(b, theta1, longth_2))
        # 从解中提取第一个解（最靠近初始猜测解的）res，并转换为以 pi 为单位的值
        res = solutions_2[0]
        theta_value_2 = solutions_2[0] / np.pi

        if theta_value_2 > 32:
            # 使之后全为0
            set_to_zero_slice(x_values, j + 2, 223)
            set_to_zero_slice(y_values, j + 2, 223)
            set_to_zero_slice(v_values, j + 2, 223)
            set_to_zero_slice(theta_values, j + 2, 223)
            break
        # 求速度
        x1 = get_x(theta1)
        y1 = get_y(theta1)
        x2 = get_x(res)
        y2 = get_y(res)
        k1 = get_k(theta1)
        k2 = get_k(res)
        k12 = (y2 - y1) / (x2 - x1)
        cos_1 = get_cos_alpha(k1, k12)
        cos_2 = get_cos_alpha(k2, k12)
        v2 = v1 * cos_1 / cos_2
        #     碰撞检测模块
        k_Ak_Ak_1=(y2-y1)/(x2-x1)

        alpha_k_k_1=get_alpha(k_Ak_Ak_1)

        c_x_3=x1+0.275*math.cos(alpha_k_k_1)+0.15*math.sin(alpha_k_k_1)
        c_y_3=y1+0.275*math.sin(alpha_k_k_1)-0.15*math.cos(alpha_k_k_1)
        c_x_4=x2-0.275*math.cos(alpha_k_k_1)+0.15*math.sin(alpha_k_k_1)
        c_y_4=y2-0.275*math.sin(alpha_k_k_1)-0.15*math.cos(alpha_k_k_1)

        # ##碰撞检测核心
        a_1 = c_x_2 - c_x_1
        a_2 = c_x_4 - c_x_3
        a_3 = c_x_3 - c_x_1
        b_1 = c_y_2 - c_y_1
        b_2 = c_y_4 - c_y_3
        b_3 = c_y_3 - c_y_1
        t=(a_3*b_2 - a_2*b_3)/(a_1*b_2 - a_2*b_1)
        s=(a_3*b_1 - a_1*b_3)/(a_1*b_2 - a_2*b_1)

        ###此处注意除了要通过碰撞条件检测碰撞与否外，
        ###还要排除碰撞第一节龙身都的情况
        if 0 <= s and s <= 1 and 0 <= t and t <= 1 and j>0:
            # 碰撞发生
            flag=1
            # 此处num为 龙身号 即第二个板凳为 龙身1
            num=j+1
        x_values.append(x2)
        y_values.append(y2)
        v_values.append(v2)
        theta_values.append(theta_value_2)
        # 更新 theta1 以供下一次迭代使用
        theta1 = res

    # 若碰撞
    # 应该存下当前 时刻
    # 输出一个碰撞到的板凳 号
    # 然后输出此时的所有板凳 x,y,v
    if flag==1:
        #碰撞时间
        print(f"The time is: {(i+1)*foot+pre_t}s")
        res_t=(i+1)*foot+pre_t
        #碰到了第几节龙身
        print(f"The collision happens at: {num}th")
        break
if flag==0:
    print("No collision happens")

row = 2  # 用于数据输入execl的行参数
# 将列表内容写入指定的行列
for m in range(0, 224):
    # 指定行列位置
    col = 2  # 列
    sheet_1.cell(row=row, column=col, value=x_values[m].item())
    col += 1  # 列+1
    sheet_1.cell(row=row, column=col, value=y_values[m].item())
    row += 1
# 指定行列位置
row = 2  # 行
col = 4  # 列
for n,v_s in enumerate(v_values):
    if isinstance(v_s, np.ndarray):
        v_s = v_s.item()  # 将NumPy数组中的单个元素转换为Python标准类型
    sheet_1.cell(row=row, column=col, value=v_s)
    row += 1
# 保存文件
workbook.save(r'F:\res_test\result2.xlsx')