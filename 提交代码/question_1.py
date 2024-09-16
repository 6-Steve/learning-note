import math
import numpy as np
from scipy.optimize import fsolve
import openpyxl

# 参数以 米 为单位
theta0 = 32 * np.pi
v1 = 1.0
b = 0.55 / (2 * np.pi)
longth = 2.86
longth_2 = 1.65

# 打开Excel文件
workbook = openpyxl.load_workbook(r'F:\res_test\result1.xlsx')
sheet_1 = workbook['位置']
sheet_2 = workbook['速度']
#方程；由theta0 求 theta1
def equation_1(theta1, b, theta0, L):
    term1 = 0.5 * theta0 * np.sqrt(1 + theta0 ** 2) + 0.5 * np.log(theta0 + np.sqrt(1 + theta0 ** 2))
    term2 = 0.5 * theta1 * np.sqrt(1 + theta1 ** 2) + 0.5 * np.log(theta1 + np.sqrt(1 + theta1 ** 2))
    return b * term1 - b * term2 - L
#切片赋零
def set_to_zero_slice(lst, x, y):
    lst[x:y + 1] = [0] * (y - x + 1)  # 使用切片赋值，y-x+1计算切片长度
    return lst

def get_x(theta):
    return b * theta * math.cos(theta)

def get_y(theta):
    return b * theta * math.sin(theta)

def get_k(theta):
    return (math.sin(theta) + theta * math.cos(theta)) / (math.cos(theta) - theta * math.sin(theta))

def get_cos_alpha(k, k12):
    # 其中k为k1或k2,返回cosα 1或2
    return abs((1 + k * k12) / math.sqrt((1 + k ** 2) * (1 + k12 ** 2)))

# 求解龙头把手后的第一个把手(求解theta)
def equation(theta, b, theta1, longth):
    # theta1为前一个，theta为当前所求
    return b ** 2 * (theta1 ** 2 + theta ** 2 - 2 * theta1 * theta * (
                math.cos(theta1) * math.cos(theta) + math.sin(theta1) * math.sin(theta))) - longth ** 2

# 初始化一个列表来存储解
solutions = []
# 循环L从1到300
# 求的是L从1到300时  龙头第一个把手的theta
for L_value in range(1, 301):
    # 初始猜测值（可能需要根据实际情况调整）
    initial_guess = 20 * np.pi  #初始猜测值
    # 使用fsolve求解
    try:
        solution = fsolve(equation_1, initial_guess, args=(b, theta0, L_value))
        if np.all(np.isfinite(solution)):  # 确保解是有限数
            solutions.append(solution[0])  # 将解添加到列表中
        else:
            solutions.append(np.nan)  # 如果解不是有限数，则记录为NaN
    except Exception as e:
        print(f"在L={L_value}时求解失败: {e}")
        solutions.append(np.nan)  # 捕获异常并记录为NaN
# 以π为单位输出解
solutions_in_pi = [sol / np.pi for sol in solutions]
# print(len(solutions_in_pi)) 输出可得大小为300
# 故从编号为0到299
for i in range(0, 300):
    # 初始化阶段：
    # 龙头第一个把手的theta
    theta1 = solutions_in_pi[i] * np.pi
    # 初始化列表来保存结果  (并且放这里起到 清零的作用)
    x_values = []
    y_values = []
    v_values = []
    theta_values = []

    initial_guess = theta1 + 0.5 * np.pi  # 初始猜测值，可以根据具体情况调整
    # 求解
    solutions = fsolve(equation, initial_guess, args=(b, theta1, longth))
    # 只要第一个解（最靠近初始猜测解的）
    res = solutions[0]
    # 输出(pi为为单位)
    theta_value = res / np.pi
    # 求位置
    x1 = get_x(theta1)
    y1 = get_y(theta1)
    v1 = 1.0
    # 求速度
    if theta_value < 32:
        x2 = get_x(res)
        y2 = get_y(res)
        k1 = get_k(theta1)
        k2 = get_k(res)
        k12 = (y2 - y1) / (x2 - x1)
        cos_1 = get_cos_alpha(k1, k12)
        cos_2 = get_cos_alpha(k2, k12)
        v2 = v1 * cos_1 / cos_2
    else:
        x2 = 0.0
        y2 = 0.0
        v2 = 0.0
        theta_value = 0.0
    x_values.append(x1)
    y_values.append(y1)
    v_values.append(v1)

    x_values.append(x2)
    y_values.append(y2)
    v_values.append(v2)
    theta_values.append(theta_value)

    theta1 = res
    for j in range(222):
        if theta1 == 0.0:#判断后面的龙身是否开始盘入
            # 若还没用盘入 使之参数全为0
            set_to_zero_slice(x_values, j + 2, 224)
            set_to_zero_slice(y_values, j + 2, 224)
            set_to_zero_slice(v_values, j + 2, 224)
            set_to_zero_slice(theta_values, j + 2, 224)
            break

        # 计算初始猜测值
        initial_guess = theta1 + 0.5 * np.pi
        solutions_2 = fsolve(equation, initial_guess, args=(b, theta1, longth_2))
        # 从解中提取第一个解（最靠近初始猜测解的）res，并转换为以 pi 为单位的值 theta_value_2
        res = solutions_2[0]
        theta_value_2 = solutions_2[0] / np.pi

        if theta_value_2 > 32:#（即大于32π）
            #判断得出还未开始盘入
            # 使之后全为0
            set_to_zero_slice(x_values, j + 2, 224)
            set_to_zero_slice(y_values, j + 2, 224)
            set_to_zero_slice(v_values, j + 2, 224)
            set_to_zero_slice(theta_values, j + 2, 224)
            break
        # 求速度
        x1 = get_x(theta1)
        y1 = get_y(theta1)
        x2 = get_x(res)
        y2 = get_y(res)
        k1 = get_k(theta1)
        k2 = get_k(res)
        k12 = (y2 - y1) / (x2 - x1)
        cos_1 = get_cos_alpha(k1, k12)
        cos_2 = get_cos_alpha(k2, k12)
        v2 = v1 * cos_1 / cos_2

        x_values.append(x2)
        y_values.append(y2)
        v_values.append(v2)
        theta_values.append(theta_value_2)
        # 更新 theta1 以供下一次迭代使用
        theta1 = res
    #将数据存入文件中
    # 指定行列位置
    row = 2  # 行
    col = i + 3  # 列
    # 将列表内容写入指定的行列
    for m in range(0, 224):
        sheet_1.cell(row=row, column=col, value=x_values[m])
        row += 1  # 行+1
        sheet_1.cell(row=row, column=col, value=y_values[m])
        row += 1
    # 指定行列位置
    row = 2  # 行
    col = i + 3  # 列
    for n in range(0, 224):
        sheet_2.cell(row=row, column=col, value=v_values[n])
        row += 1
# 保存文件
workbook.save(r'F:\res_test\result1.xlsx')