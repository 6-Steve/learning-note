import numpy as np
import math
from scipy.optimize import fsolve
import openpyxl

workbook = openpyxl.load_workbook(r'F:\res_test\result4.xlsx')

sheet_1 = workbook['位置']
sheet_2 = workbook['速度']

# 在螺线中已知theta0，长度L(曲线长度)，螺距b，求theta1
# theta0对应未知theta(大)，theta1对应已知theta_C1(小)
def equation_1(theta0, b, theta1, L):
    term1 = 0.5 * theta0 * np.sqrt(1 + theta0 ** 2) + 0.5 * np.log(theta0 + np.sqrt(1 + theta0 ** 2))
    term2 = 0.5 * theta1 * np.sqrt(1 + theta1 ** 2) + 0.5 * np.log(theta1 + np.sqrt(1 + theta1 ** 2))
    return b * term1 - b * term2 - L

#计算螺线上的后把手theta
def equation(theta, b, theta1, longth):
    # theta1为前一个，theta为当前所求
    return b ** 2 * (theta1 ** 2 + theta ** 2 - 2 * theta1 * theta * (
                math.cos(theta1) * math.cos(theta) + math.sin(theta1) * math.sin(theta))) - longth ** 2
#不同圆 仅x_o，y_o，r值不同,然后x_1,y_1每次都不同
def get_theta_in_o(theta, x_o , y_o , r_ , x_1 , y_1 ,longth):  
    term1 = r_**2 + (x_o - x_1)**2 + (y_o - y_1)**2
    term2 = 2*(x_o - x_1)*r_*math.cos(theta) + 2*(y_o - y_1)*r_*math.sin(theta)
    return term1 + term2 - longth**2

def pre_in_o_las_in_line(theta,b,x_1,y_1,longth):#前点在圆，后点在螺线上
    term1 = x_1**2 + y_1**2 + (b**2)*(theta**2)
    term2 = 2 * x_1 * b * theta * math.cos(theta) + 2 * y_1 * b * theta * math.sin(theta)
    return term1 - term2 - longth**2

#得到下一个把手的theta和type
def get_next(point_theta,point_type,longth):
    # 应该返回下一个点的theta和type
    #根据分类 判断位置
    if point_type == 1:  #龙头前把手在盘入螺线上
        #后把手均在盘入螺线上
        #求解在盘入螺线上的theta
        theta1 = point_theta
        initial_guess = theta1 + 0.5 * np.pi  # 初始猜测值，可以根据具体情况调整
        # 求解
        solve_1 = fsolve(equation, initial_guess, args=(b, theta1, longth))
        # 只要第一个解（最靠近初始猜测解的）
        res_1 = solve_1[0]
        # 保存结果
        return res_1,1  # 1代表龙头后把手在盘入螺线上

    elif point_type == 2:  #龙头前把手在圆o1上
        #后把手可能在圆o1上 也 可能在盘入螺线上
        #通过 龙头前把手和C1的距离 判断

        #首先得到龙头前把手的笛卡尔坐标
        x_2 = o_1_x + r1 * np.cos(point_theta)
        y_2 = o_1_y + r1 * np.sin(point_theta)
        #判断
        dist=np.sqrt( (x_2-x_C1)**2 + (y_2-y_C1)**2 )
        if dist > longth: #即后把手在圆o1上
            #后把手在圆o1上
            # 定义初始猜测值
            initial_guess = point_theta + 0.25 * np.pi  
            # 使用fsolve求解
            solve_2 = fsolve(get_theta_in_o, initial_guess, args=(o_1_x,o_1_y,r1,x_2,y_2,longth))
            res_2 = solve_2[0]
            # 保存结果
            return res_2 ,2  # 2代表龙头后把手在圆o1上
        else:
            #后把手在盘入螺线上
            #求解在盘入螺线上的theta
            theta1 = point_theta
            initial_guess = theta_C1 + 0.5 * np.pi  
            # 求解
            solve_2 = fsolve(pre_in_o_las_in_line, initial_guess, args=(b, x_2,y_2, longth)) 
            # 只要第一个解（最靠近初始猜测解的）
            res_2 = solve_2[0]
            # 保存结果
            return res_2 ,1  # 1代表龙头后把手在盘入螺线上

    elif point_type == 3:  #龙头前把手在圆o2上
        #后把手可能在圆o2上 也 可能在圆o1上
        #通过 龙头前把手和C3的距离 判断
        #首先得到龙头前把手的笛卡尔坐标
        x_3 = o_2_x + r2 * np.cos(point_theta)
        y_3 = o_2_y + r2 * np.sin(point_theta)
        #判断
        dist=np.sqrt( (x_3-x_C3)**2 + (y_3-y_C3)**2 )
        if dist > longth: #即后把手在圆o2上
            #后把手在圆o2上
            #get_theta_in_o(theta, x_o , y_o , x_1 , y_1 ,longth)
            # 定义初始猜测值
            initial_guess = point_theta - 0.6 * np.pi  
            # 使用fsolve求解
            solve_3 = fsolve(get_theta_in_o, initial_guess, args=(o_2_x, o_2_y, r2,x_3, y_3, longth))
            res_3 = solve_3[0]
            # 保存结果
            return res_3 ,3  # 3代表龙头后把手在圆o2上
        else:
            #后把手在圆o1上
            # 定义初始猜测值
            initial_guess = theta_C3 + 0.25 * np.pi  
            # 使用fsolve求解
            solve_3 = fsolve(get_theta_in_o, initial_guess, args=(o_1_x, o_1_y, r1,x_3, y_3, longth))
            res_3 = solve_3[0]
            # 保存结果
            return res_3 ,2  # 2代表龙头后把手在圆o1上

    else:  #龙头前把手在盘出螺线上
        #后把手可能在盘出螺线上 也 可能在圆o2上
        #通过 龙头前把手和C2的距离 判断
        #首先得到龙头前把手的笛卡尔坐标
        x_4 = -b*point_theta*np.cos(point_theta)
        y_4 = -b*point_theta*np.sin(point_theta)
        #判断
        dist=np.sqrt( (x_4-x_C2)**2 + (y_4-y_C2)**2 ) 
        if dist >= longth or point_theta > 6*np.pi: #即后把手在盘出螺线上
            #后把手在盘出螺线上
            # 定义初始猜测值
            theta_4 = point_theta
            initial_guess = theta_4 - 0.5 * np.pi  
            # 使用fsolve求解
            solve_4 = fsolve(equation, initial_guess, args=(b, theta_4, longth))
            res_4 = solve_4[0]
            # 保存结果
            return res_4 ,4  # 4代表龙头后把手在盘出螺线上
        else:
            #后把手在圆o2上
            # 定义初始猜测值
            initial_guess = 3.75*np.pi  
            # 使用fsolve求解
            solve_4 = fsolve(get_theta_in_o, initial_guess, args=(o_2_x, o_2_y, r2,x_4, y_4, longth))
            res_4 = solve_4[0]
            # 保存结果
            return res_4 ,3  # 3代表龙头后把手在圆o2上
def get_k(theta_,type_):
    if type_ == 1 or type_ == 4:
        k=( math.sin(theta_)+theta_*math.cos(theta_) ) / ( math.cos(theta_)-theta_*math.sin(theta_) )
    else:   # type_ == 2 or type_ == 3
        k=-math.cos(theta_)/math.sin(theta_)
    return k
def get_x_y(p_theta,p_theta_type):
    if p_theta_type == 1:#盘入螺线
            x = b*p_theta*np.cos(p_theta)
            y = b*p_theta*np.sin(p_theta)
    elif p_theta_type == 2:#圆o1
        x = o_1_x + r1 * np.cos(p_theta)
        y = o_1_y + r1 * np.sin(p_theta)
    elif p_theta_type == 3:#圆o2
        x = o_2_x + r2 * np.cos(p_theta)
        y = o_2_y + r2 * np.sin(p_theta)
    else: #盘出螺线
        x = -b*p_theta*np.cos(p_theta)
        y = -b*p_theta*np.sin(p_theta)
    return x,y

# 求速度函数
def get_v(theta1,type1,theta2,type2,x1,y1,v1,x2,y2):
    k1=get_k(theta1,type1)
    k2=get_k(theta2,type2)
    k12=(y2-y1)/(x2-x1)
    cos_alpha1 = abs((1 + k1 * k12) / math.sqrt((1 + k1 ** 2) * (1 + k12 ** 2)))
    cos_alpha2 = abs((1 + k2 * k12) / math.sqrt((1 + k2 ** 2) * (1 + k12 ** 2)))
    v2=v1*cos_alpha1/cos_alpha2
    return v2

# 均为弧度制
#初始化已知量（单位为m） 分别为 螺距b 与 掉头空间半径r
b=1.7 / (2 * np.pi)
r=4.5
longth = 2.86
longth_2 = 1.65
#求点C1,C2坐标
theta_C1= r / b
theta_C2= theta_C1
x_C1= b * theta_C1 * math.cos(theta_C1)
y_C1= b * theta_C1 * math.sin(theta_C1)

x_C2= -b * theta_C2 * math.cos(theta_C2)
y_C2= -b * theta_C2 * math.sin(theta_C2)

#求k1，k2
k2=( math.sin(theta_C1)+theta_C1*math.cos(theta_C1) )/( math.cos(theta_C1)-theta_C1*math.sin(theta_C1) )
k1=-x_C1 / y_C1
cos_beta= (1 + k1 * k2) / math.sqrt((1 + k1 ** 2) * (1 + k2 ** 2)) 
beta=math.acos(cos_beta)
alpha=np.pi/2-beta
#求r1，r2
r2= r/(3*math.sin(alpha))
r1= 2*r2
#求角A j_A 和（alpha_1）
alpha_1=math.atan(k1) + np.pi
j_A=-alpha_1
#求圆心o_1，o_2
o_1_x = x_C1/3 - r*math.tan(beta)*math.cos(alpha_1)*2/3
o_1_y = y_C1/3 - r*math.tan(beta)*math.sin(alpha_1)*2/3

o_2_x = -x_C1*2/3 + r*math.tan(beta)*math.cos(alpha_1)/3
o_2_y = -y_C1*2/3 + r*math.tan(beta)*math.sin(alpha_1)/3
#求C3坐标
x_C3 = o_1_x/3 + o_2_x*2/3
y_C3 = o_1_y/3 + o_2_y*2/3

# C3在圆o1上的theta
theta_C3 = alpha - j_A - (2*r1*alpha)/r1 

##求 龙头把手  位置 （分四种情况）
# 1. 龙头把手在盘入螺线上  -100<= t <=0时
list_head_1 = []
for L_path in range(-100, 1):
    #L_path在变化（速度为1，则为从-100到0）
    #取绝对值
    L_path = abs(L_path)
    #求解
    #猜测初始值
    initial_guess = theta_C1-np.pi  
    #求解
    solution_1 = fsolve(equation_1, initial_guess, args=(b, theta_C1, L_path))
    #将结果添加到列表
    list_head_1.append(solution_1)
#双list，此list表示对应   类type  （注意 i 应一一对应）
list_head_1_type = [1] * len(list_head_1)  # 1代表龙头把手在盘入螺线上

# 2. 龙头把手在圆o1上  0< t <= 2*r1*alpha 时
list_head_2 = []
max_1=math.floor(2*r1*alpha)  #max_1为 9
# 2*r1*alpha值为 9.080829937880747，应该向下取整，即9s时在圆o1上(差点出去)，10s时在圆o2上
for L_path in range(1, max_1+1):
    j_a=L_path/r1
    theta_o_1= alpha -  j_A -j_a
    #算出theta_o_1保存就好，暂时不用求对应x,y
    list_head_2.append(theta_o_1)
list_head_2_type = [2] * len(list_head_2) 
# 3. 龙头把手在圆o2上  2*r1*alpha < t <= 2*r1*alpha + 2*r2*alpha时
list_head_3 = []
max_2=math.floor(2*r1*alpha+2*r2*alpha)  #max_2为 13
# 2*r1*alpha + 2*r2*alpha+1值为 13.62124490682112，应该向下取整，即13s时在圆o2上，14s时在盘出螺线上
for L_path in range(max_1+1, max_2+1):
    j_b=( L_path - 2*alpha*r1 ) / r2
    theta_o_2= np.pi - j_A - alpha + j_b
    #算出theta_o_2保存就好，暂时不用求对应x,y
    list_head_3.append(theta_o_2)
list_head_3_type = [3] * len(list_head_3)  

# 4. 龙头把手在盘出螺线上  2*r1*alpha + 2*r2*alpha < t <= 100时
list_head_4 = []
for L_path in range(max_2+1, 101):
    #L_path在变化（速度为1，则为从14到100）
    # 在盘出螺线上的实际弧长值
    L_path_real=L_path-2*r1*alpha-2*r2*alpha
    #猜测初始值
    initial_guess = theta_C2+np.pi  
    #求解
    # ！此处b不用改为-b，因为公式推导时把b**2从根号下提出来后相当于 取了绝对值
    solution_4 = fsolve(equation_1, initial_guess, args=(b, theta_C2, L_path_real))
    #将结果添加到列表
    list_head_4.append(solution_4)
list_head_4_type = [4] * len(list_head_4)  
##龙头前把手-100s 到 100s 位置 求解完毕

###下面进行推导后续把手工作 分为 龙头后把手 和 龙身把手
##合并所有龙头前把手位置 和 对应类的  list 
mix_list_head = list_head_1 + list_head_2 + list_head_3 + list_head_4
mix_list_head_type = list_head_1_type + list_head_2_type + list_head_3_type + list_head_4_type
# 初始化存龙头后把手的list 和 记录类型的list
mix_list_head_2 = []
mix_list_head_2_type = []
#遍历所有龙头前把手位置，根据分类 来分情况处理
for i in range(len(mix_list_head)): #0到200 正好对应1到201个
    theta_next_0,type_next_0 = get_next(mix_list_head[i],mix_list_head_type[i],longth)
    mix_list_head_2.append(theta_next_0)
    mix_list_head_2_type.append(type_next_0)

#龙身把手
#遍历所有龙头后把手位置，根据分类 来分情况处理
#每一个龙头后把手位置对应一个时刻（秒）
#一个循环为一秒，在循环里求这个时刻 其后所有把手theta，类 存为两个list 
#在循环结束时根据 theta 和 类 转换为 坐标， 并 求出速度 输出
for i in range(len(mix_list_head_2)):  #len(mix_list_head_2) 0~200  (201)
    #初始化存储list
    list_point_theta = []
    list_point_theta_type = []
    #在其中加入龙头 前把手 和 后把手
    list_point_theta.append(mix_list_head[i])
    list_point_theta.append(mix_list_head_2[i])
    list_point_theta_type.append(mix_list_head_type[i])
    list_point_theta_type.append(mix_list_head_2_type[i])
    #当前时刻的龙身把手theta(为第二个把手，设为theta02)
    theta_pre = mix_list_head_2[i]
    #当前时刻的龙头后把手类型
    type_pre = mix_list_head_2_type[i]
    #要得到所有剩下的222个把手theta和类（共有224 224-2=222）
    for j in range(0,222): #0~221
        #带入函数计算
        theta_next,type_next = get_next(theta_pre,type_pre,longth_2)
        #存储结果
        list_point_theta.append(theta_next)
        list_point_theta_type.append(type_next)
        #更新上一个把手的值
        theta_pre=theta_next
        type_pre=type_next
    #已经存储完毕224个把手的theta和类
    #根据theta和类 转换为 坐标， 并 求出速度 输出
    #初始化存储数组
    list_x = []
    list_y = []
    list_v = []
    for k in range(len(list_point_theta)):
        #根据类 判断 所处 阶段（1，2，3，4）
        p_theta=list_point_theta[k]
        p_theta_type=list_point_theta_type[k]

        x,y=get_x_y(p_theta,p_theta_type)
        list_x.append(x)
        list_y.append(y)
        #求速度
        if k == 0:
            v = 1.0
        else:
            v=get_v(list_point_theta[k-1],list_point_theta_type[k-1],
                    p_theta,p_theta_type,
                    list_x[k-1],list_y[k-1],list_v[k-1],
                    x,y)
        list_v.append(v)
    #将文件输入execl中
    # 指定行列位置
    row = 2  # 行
    col = i + 2  # 列
    # 将列表内容写入指定的行列
    for m in range(0, 224):
        sheet_1.cell(row=row, column=col, value=list_x[m].item())
        row += 1  # 行+1
        sheet_1.cell(row=row, column=col, value=list_y[m].item())
        row += 1
    # 指定行列位置
    row = 2  # 行
    col = i + 2  # 列
    for n,v_s in enumerate(list_v):
        if isinstance(v_s, np.ndarray):
            v_s = v_s.item()  # 将NumPy数组中的单个元素转换为Python标准类型
        sheet_2.cell(row=row, column=col, value=v_s)
        row += 1
# 保存文件
workbook.save(r'F:\res_test\result4.xlsx')